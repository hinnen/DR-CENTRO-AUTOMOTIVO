"""Utilitários comuns para importação/exportação Excel."""

from __future__ import annotations

import csv
import unicodedata
from io import BytesIO
from pathlib import Path
from typing import Any
from uuid import UUID

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill

IMPORT_MAX_ROWS = 5000
HEADER_FILL = PatternFill("solid", fgColor="E8EEF9")
HEADER_FONT = Font(bold=True, color="152E69")

BOOL_TRUE = frozenset({"1", "sim", "s", "true", "verdadeiro", "ativo", "yes", "y"})
BOOL_FALSE = frozenset({"0", "nao", "não", "n", "false", "falso", "inativo", "no"})


def normalize_header(value: str) -> str:
    text = unicodedata.normalize("NFD", str(value or ""))
    text = "".join(c for c in text if unicodedata.category(c) != "Mn")
    return text.strip().lower()


def parse_bool(value: Any, *, default: bool | None = None) -> bool | None:
    if value is None or str(value).strip() == "":
        return default
    token = normalize_header(str(value))
    if token in BOOL_TRUE:
        return True
    if token in BOOL_FALSE:
        return False
    return default


def cell_str(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def parse_uuid(value: Any) -> UUID | None:
    raw = cell_str(value)
    if not raw:
        return None
    try:
        return UUID(raw)
    except ValueError:
        return None


def read_spreadsheet(upload_path: Path) -> dict[str, list[dict[str, Any]]]:
    """Retorna {nome_aba: [linhas dict por cabeçalho original]}."""
    suffix = upload_path.suffix.lower()
    if suffix == ".csv":
        return {"Cadastro": _read_csv(upload_path)}
    if suffix in {".xlsx", ".xls"}:
        return _read_xlsx(upload_path)
    raise ValueError("Use arquivo .csv ou .xlsx.")


def _read_csv(path: Path) -> list[dict[str, Any]]:
    for encoding in ("utf-8-sig", "latin-1", "cp1252"):
        try:
            with path.open("r", encoding=encoding, newline="") as handle:
                reader = csv.DictReader(handle, delimiter=";")
                if reader.fieldnames and len(reader.fieldnames) == 1:
                    handle.seek(0)
                    reader = csv.DictReader(handle, delimiter=",")
                return [dict(row) for row in reader]
        except UnicodeDecodeError:
            continue
    raise ValueError("Não foi possível ler o CSV (encoding).")


def _read_xlsx(path: Path) -> dict[str, list[dict[str, Any]]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheets: dict[str, list[dict[str, Any]]] = {}
    for worksheet in workbook.worksheets:
        iterator = worksheet.iter_rows(values_only=True)
        headers = [cell_str(cell) for cell in next(iterator, [])]
        if not any(headers):
            continue
        rows: list[dict[str, Any]] = []
        for row in iterator:
            if not any(row):
                continue
            item = {headers[index]: row[index] if index < len(row) else None for index, header in enumerate(headers) if header}
            rows.append(item)
        sheets[worksheet.title] = rows
    workbook.close()
    return sheets


def build_workbook(*, sheets: list[tuple[str, list[tuple[str, str]], list[dict[str, Any]]]]) -> bytes:
    """Monta XLSX com abas (título, [(rótulo, chave)], linhas)."""
    workbook = Workbook()
    workbook.remove(workbook.active)
    for title, headers, rows in sheets:
        worksheet = workbook.create_sheet(title=title[:31])
        for column, (label, _key) in enumerate(headers, start=1):
            cell = worksheet.cell(row=1, column=column, value=label)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
        for row_index, row in enumerate(rows, start=2):
            for column, (_label, key) in enumerate(headers, start=1):
                worksheet.cell(row=row_index, column=column, value=row.get(key, ""))
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def map_rows_by_header(rows: list[dict[str, Any]], header_map: dict[str, str]) -> list[dict[str, Any]]:
    """Converte cabeçalhos da planilha (rótulo ou chave) para chaves internas."""
    normalized_map = {normalize_header(label): key for label, key in header_map.items()}
    mapped: list[dict[str, Any]] = []
    for row in rows[:IMPORT_MAX_ROWS]:
        item: dict[str, Any] = {}
        for header, value in row.items():
            key = normalized_map.get(normalize_header(header))
            if key:
                item[key] = value
        if any(str(value).strip() for value in item.values()):
            mapped.append(item)
    return mapped
