"""Testes de dados de exemplo e planilhas."""

from io import BytesIO

from django.contrib.auth import authenticate
from django.core.exceptions import ValidationError
from django.test import TestCase
from django.urls import reverse
from openpyxl import load_workbook

from apps.accounts.models import Role
from apps.accounts.tests import make_user
from apps.core.models import WorkshopSettings
from apps.core.services.demo_purge import purge_demo_data
from apps.core.services.demo_seed import DemoDataAlreadyLoaded, load_demo_data
from apps.core.services.settings import invalidate_workshop_settings_cache
from apps.core.spreadsheet import build_export_workbook, import_uploaded_file
from apps.core.spreadsheet.sheets import SHEET_CLIENTS, SHEET_LOCATIONS, SHEET_USERS, SHEET_VEHICLES
from apps.customers.models import Client
from apps.vehicles.models import Vehicle, VehicleLocation
from apps.workorders.models import ServiceOrder, Status


class DemoDataTests(TestCase):
    ADMIN_PASSWORD = "senha-forte-123"

    def setUp(self):
        self.admin = make_user("admin_demo", Role.ADMIN)
        WorkshopSettings.load()
        invalidate_workshop_settings_cache()

    def test_load_demo_creates_orders_in_multiple_statuses(self):
        counts = load_demo_data(actor=self.admin)
        self.assertGreaterEqual(counts["orders"], 10)
        self.assertTrue(ServiceOrder.objects.filter(is_demo=True).exists())
        statuses = set(ServiceOrder.objects.filter(is_demo=True).values_list("status", flat=True))
        self.assertIn(Status.IN_SERVICE, statuses)
        self.assertIn(Status.WAITING_EVALUATION, statuses)

    def test_load_demo_twice_raises(self):
        load_demo_data(actor=self.admin)
        with self.assertRaises(DemoDataAlreadyLoaded):
            load_demo_data(actor=self.admin)

    def test_purge_requires_admin_password(self):
        load_demo_data(actor=self.admin)
        with self.assertRaises(ValidationError):
            purge_demo_data(actor=self.admin, password="errada")
        self.assertTrue(ServiceOrder.objects.filter(is_demo=True).exists())

    def test_purge_removes_demo_records(self):
        load_demo_data(actor=self.admin)
        purge_demo_data(actor=self.admin, password=self.ADMIN_PASSWORD)
        self.assertEqual(ServiceOrder.objects.filter(is_demo=True).count(), 0)
        self.assertEqual(Vehicle.objects.filter(is_demo=True).count(), 0)
        self.assertEqual(Client.objects.filter(is_demo=True).count(), 0)

    def test_demo_views_require_admin(self):
        reception = make_user("recep_demo", Role.RECEPTION)
        self.client.force_login(reception)
        response = self.client.get(reverse("core:settings_demo"))
        self.assertEqual(response.status_code, 403)


class SpreadsheetTests(TestCase):
    def setUp(self):
        self.admin = make_user("admin_sheet", Role.ADMIN)

    def _upload_bytes(self, blob: bytes, name: str = "cadastro.xlsx"):
        from django.core.files.uploadedfile import SimpleUploadedFile

        upload = SimpleUploadedFile(
            name,
            blob,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        return import_uploaded_file(upload=upload, actor=self.admin)

    def test_export_workbook_has_four_sheets(self):
        blob = build_export_workbook(mode="template", actor=self.admin)
        workbook = load_workbook(BytesIO(blob), read_only=True)
        self.assertEqual(
            set(workbook.sheetnames),
            {SHEET_LOCATIONS, SHEET_CLIENTS, SHEET_VEHICLES, SHEET_USERS},
        )
        workbook.close()

    def test_import_client_and_location_roundtrip(self):
        blob = build_export_workbook(mode="template", actor=self.admin)
        summary = self._upload_bytes(blob)
        self.assertEqual(summary.clients.skipped, 1)
        self.assertEqual(summary.locations.skipped, 1)

        workbook = load_workbook(BytesIO(blob))
        loc = workbook[SHEET_LOCATIONS]
        loc.cell(row=2, column=2, value="Box Teste")
        loc.cell(row=2, column=1, value="")
        cli = workbook[SHEET_CLIENTS]
        cli.cell(row=2, column=2, value="Cliente Planilha")
        cli.cell(row=2, column=3, value="13988776655")
        cli.cell(row=2, column=7, value="")
        cli.cell(row=2, column=1, value="")
        loc.cell(row=2, column=4, value="Sim")
        buffer = BytesIO()
        workbook.save(buffer)
        workbook.close()

        summary = self._upload_bytes(buffer.getvalue())
        self.assertEqual(summary.locations.created, 1)
        self.assertEqual(summary.clients.created, 1)
        self.assertTrue(VehicleLocation.objects.filter(name="Box Teste").exists())
        self.assertTrue(Client.objects.filter(phone="13988776655").exists())

    def test_import_user_requires_pin_for_new(self):
        blob = build_export_workbook(mode="template", actor=self.admin)
        workbook = load_workbook(BytesIO(blob))
        users = workbook[SHEET_USERS]
        users.cell(row=2, column=2, value="Mec Planilha")
        users.cell(row=2, column=3, value="mec_plan")
        users.cell(row=2, column=4, value="Mecânico")
        users.cell(row=2, column=6, value="5566")
        users.cell(row=2, column=1, value="")
        buffer = BytesIO()
        workbook.save(buffer)
        workbook.close()

        summary = self._upload_bytes(buffer.getvalue())
        self.assertEqual(summary.users.created, 1)
        self.assertEqual(authenticate(username="mec_plan", password="5566").role, Role.MECHANIC)

    def test_export_locations_uses_pk_not_uuid(self):
        VehicleLocation.objects.create(name="Box Export", order=1)
        blob = build_export_workbook(mode="export", actor=self.admin)
        workbook = load_workbook(BytesIO(blob), read_only=True)
        sheet = workbook[SHEET_LOCATIONS]
        rows = list(sheet.iter_rows(min_row=2, values_only=True))
        workbook.close()
        self.assertTrue(any(row[1] == "Box Export" for row in rows if row))

    def test_spreadsheet_download_requires_admin(self):
        reception = make_user("recep_sheet", Role.RECEPTION)
        self.client.force_login(reception)
        response = self.client.get(reverse("core:settings_spreadsheet_download"))
        self.assertEqual(response.status_code, 403)
